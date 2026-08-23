from __future__ import annotations

import argparse
import json
import math
import statistics
from collections import defaultdict
from pathlib import Path
from tempfile import TemporaryDirectory
from zipfile import ZipFile

from openpyxl import load_workbook


def percentile(values: list[float], probability: float) -> float:
    ordered = sorted(values)
    position = (len(ordered) - 1) * probability
    lower = math.floor(position)
    upper = math.ceil(position)
    if lower == upper:
        return ordered[lower]
    return ordered[lower] * (upper - position) + ordered[upper] * (position - lower)


def find_workbook(source: Path, temporary_directory: Path) -> Path:
    if source.suffix.lower() == ".xlsx":
        return source
    if source.suffix.lower() != ".zip":
        raise ValueError("Source must be the Online Retail II ZIP archive or XLSX workbook")
    with ZipFile(source) as archive:
        candidates = [name for name in archive.namelist() if name.lower().endswith(".xlsx")]
        if len(candidates) != 1:
            raise ValueError("The archive must contain exactly one XLSX workbook")
        archive.extract(candidates[0], temporary_directory)
        return temporary_directory / candidates[0]


def train(source: Path) -> dict:
    with TemporaryDirectory(prefix="paymentor-retail-") as temp:
        workbook_path = find_workbook(source, Path(temp))
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        daily = defaultdict(lambda: {"sales": 0.0, "returns": 0.0})
        row_count = cancellation_count = excluded_count = 0
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for invoice, _stock, _description, quantity, invoice_date, price, _customer, _country in sheet.iter_rows(
                min_row=2, values_only=True
            ):
                row_count += 1
                if invoice_date is None or quantity is None or price is None:
                    excluded_count += 1
                    continue
                try:
                    quantity_value = float(quantity)
                    line_value = quantity_value * float(price)
                except (TypeError, ValueError):
                    excluded_count += 1
                    continue
                day = invoice_date.date()
                if str(invoice).startswith("C") or quantity_value < 0:
                    daily[day]["returns"] += abs(line_value)
                    cancellation_count += 1
                elif quantity_value > 0 and float(price) > 0:
                    daily[day]["sales"] += line_value
                else:
                    excluded_count += 1

    net_values = [record["sales"] - record["returns"] for record in daily.values()]
    lower_bound = percentile(net_values, 0.02)
    upper_bound = percentile(net_values, 0.98)
    winsorized = {
        day: min(max(record["sales"] - record["returns"], lower_bound), upper_bound)
        for day, record in daily.items()
    }
    baseline = statistics.median(winsorized.values())
    weekday_values = defaultdict(list)
    month_values = defaultdict(list)
    for day, value in winsorized.items():
        weekday_values[day.weekday()].append(value)
        month_values[day.month].append(value)
    gross_sales = sum(record["sales"] for record in daily.values())
    returns = sum(record["returns"] for record in daily.values())
    return {
        "schema_version": 1,
        "model_name": "explainable_retail_seasonality_v1",
        "source": {
            "dataset": "UCI Online Retail II",
            "file": source.name,
            "currency": "GBP",
            "first_date": str(min(daily)),
            "last_date": str(max(daily)),
            "rows_seen": row_count,
            "active_days": len(daily),
            "cancellation_lines": cancellation_count,
            "excluded_lines": excluded_count,
        },
        "training": {
            "target": "daily_net_sales",
            "winsorization": {"lower_quantile": 0.02, "upper_quantile": 0.98},
            "baseline_daily_net_sales": round(baseline, 6),
            "daily_mean": round(statistics.mean(winsorized.values()), 6),
            "daily_stddev": round(statistics.stdev(winsorized.values()), 6),
            "return_rate": round(returns / gross_sales, 8),
            "weekday_multipliers": {
                str(index): round(statistics.mean(weekday_values[index]) / baseline, 6)
                if weekday_values[index]
                else 0.0
                for index in range(7)
            },
            "month_multipliers": {
                str(month): round(statistics.mean(month_values[month]) / baseline, 6)
                if month_values[month]
                else 1.0
                for month in range(1, 13)
            },
        },
        "demo_assumptions": {
            "target_currency": "INR",
            "baseline_daily_inflow_paise": 5000000,
            "opening_balance_paise": 25000000,
            "safe_reserve_paise": 10000000,
            "variable_cost_ratio": 0.42,
            "payment_fee_ratio": 0.022,
            "fixed_daily_opex_paise": 600000,
        },
        "limitations": [
            "The source contains sales and cancellations, not bank balances or operating expenses.",
            "INR scale and expense assumptions are synthetic demo inputs, not learned facts.",
            "This artifact is a global demo prior and is never trained on a Paymentor tenant's private data.",
        ],
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Train Paymentor's explainable retail cash-flow prior")
    parser.add_argument("source", type=Path, help="Path to online+retail+ii.zip or online_retail_II.xlsx")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(__file__).resolve().parents[1] / "app" / "data" / "retail_cashflow_model.json",
    )
    args = parser.parse_args()
    artifact = train(args.source)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(artifact, indent=2) + "\n", encoding="utf-8")
    print(f"Trained {artifact['model_name']} from {artifact['source']['rows_seen']:,} rows")
    print(f"Saved model artifact to {args.output}")


if __name__ == "__main__":
    main()
